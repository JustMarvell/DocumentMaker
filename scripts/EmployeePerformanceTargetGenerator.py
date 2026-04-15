
"""
EmployeePerformanceTargetGenerator.py
Generates 'Sasaran Kinerja Pegawai' (.xlsx) by filling placeholders in the template.
 
Usage (called by DocumentController.php):
    python3 EmployeePerformanceTargetGenerator.py \
        --appraisal-period-start "2025-01-01" \
        --appraisal-period-end "2025-12-31" \
        --employee-name "John Doe" \
        --employee-nip "7103123456789" \
        --employee-rank "Penata III/c" \
        --employee-position "Staff Administratif" \
        --employee-work-unit "Unit Administratif" \
        --appraisal-name "Jane Doe" \
        --appraisal-nip "7103987654321" \
        --appraisal-rank "Penata II" \
        --appraisal-position "Kepala Unit" \
        --appraisal-work-unit "Unit Staff" \
        --leadership-work-result-plan "Rencana kerja pimpinan" \
        --work-result-plan "Rencana hasil kerja" \
        --work-quantity-indicator "Jumlah dokumen" \
        --work-quality-indicator "Ketepatan isi dokumen" \
        --work-time-indicator "Tepat waktu" \
        --work-quantity-target "12 dokumen" \
        --work-quality-target "100%" \
        --work-time-target "12 bulan" \
        --additional-work-behaviour-1 "Integritas" \
        --additional-work-behaviour-1-description "Bertindak jujur dan konsisten" \
        --leadership-spesific-expectation "Meningkatkan kualitas laporan"
"""

import argparse
import sys
import re
from pathlib import Path
from openpyxl import load_workbook

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate Employee Performance Targets (.xlsx)")
    parser.add_argument("--appraisal-period-start", required=True)
    parser.add_argument("--appraisal-period-end", required=True)
    parser.add_argument("--employee-name", required=True)
    parser.add_argument("--employee-nip", required=True)
    parser.add_argument("--employee-rank", required=True)
    parser.add_argument("--employee-position", required=True)
    parser.add_argument("--employee-work-unit", required=True)
    parser.add_argument("--appraisal-name", required=True)
    parser.add_argument("--appraisal-nip", required=True)
    parser.add_argument("--appraisal-rank", required=True)
    parser.add_argument("--appraisal-position", required=True)
    parser.add_argument("--appraisal-work-unit", required=True)
    parser.add_argument("--leadership-work-result-plan", required=True)
    parser.add_argument("--work-result-plan", required=True)
    parser.add_argument("--work-quantity-indicator", required=True)
    parser.add_argument("--work-quality-indicator", required=True)
    parser.add_argument("--work-time-indicator", required=True)
    parser.add_argument("--work-quantity-target", required=True)
    parser.add_argument("--work-quality-target", required=True)
    parser.add_argument("--work-time-target", required=True)
    parser.add_argument("--additional-work-behaviour-1", required=True)
    parser.add_argument("--additional-work-behaviour-1-description", required=True)
    parser.add_argument("--leadership-spesific-expectation", required=True)
    parser.add_argument("--output-filename",    required=True)
    return parser.parse_args()
    
def build_context(args: argparse.Namespace) -> dict:
    return {
        "appraisal_period_start" : args.appraisal_period_start,
        "appraisal_period_end" : args.appraisal_period_end,
        "employee_name" : args.employee_name,
        "employee_nip" : args.employee_nip,
        "employee_rank" : args.employee_rank,
        "employee_position" : args.employee_position,
        "employee_work_unit" : args.employee_work_unit,
        "appraisal_name" : args.appraisal_name,
        "appraisal_nip" : args.appraisal_nip,
        "appraisal_rank" : args.appraisal_rank,
        "appraisal_position" : args.appraisal_position,
        "appraisal_work_unit" : args.appraisal_work_unit,
        "leadership_work_result_plan" : args.leadership_work_result_plan,
        "work_result_plan" : args.work_result_plan,
        "work_quantity_indicator" : args.work_quantity_indicator,
        "work_quality_indicator" : args.work_quality_indicator,
        "work_time_indicator" : args.work_time_indicator,
        "work_quantity_target" : args.work_quantity_target,
        "work_quality_target" : args.work_quality_target,
        "work_time_target" : args.work_time_target,
        "additional_work_behaviour_1" : args.additional_work_behaviour_1,
        "additional_work_behaviour_1_description" : args.additional_work_behaviour_1_description,
        "leadership_spesific_expectation" : args.leadership_spesific_expectation,
    }
    
def fill_placeholders(ws, context: dict) -> None:
    placeholder_re = re.compile(r"\{\{\s*([\w]+)\s*\}\}")
    
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            original = cell.value
            
            def replacer(match):
                key = match.group(1)
                if key in context:
                    return str(context[key])
                return match.group(0)

            new_value = placeholder_re.sub(replacer, original)
            if new_value != original:
                cell.value = new_value
                
def main() -> None:
    args = parse_args()
    context = build_context(args)
 
    base_dir      = Path(__file__).resolve().parent.parent
    template_path = base_dir / "document_templates" / "SKP-Template.xlsx"
    output_dir    = base_dir / "public" / "cached_result"
    output_path   = output_dir / args.output_filename
 
    if not template_path.exists():
        print(f"ERROR: Template not found at {template_path}", file=sys.stderr)
        sys.exit(1)
 
    output_dir.mkdir(parents=True, exist_ok=True)
 
    wb = load_workbook(str(template_path))
    for sheet_name in wb.sheetnames:
        fill_placeholders(wb[sheet_name], context)
 
    wb.save(str(output_path))
    print(f"OK: Saved to {output_path}")
 
 
if __name__ == "__main__":
    main()