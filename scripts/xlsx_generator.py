"""
xlsx_generator.py
Generic Excel document generator using pure zip/XML manipulation.

Approach:
  - Never uses openpyxl.save() — which destroys drawings and text boxes
  - Works directly on the xlsx zip file contents:
      * xl/sharedStrings.xml  → render {{ }} using <t>...</t> pattern
      * xl/drawings/*.xml     → render {{ }} using <a:t>...</a:t> pattern
      * Everything else       → copied verbatim (sheet, styles, images,
                                relationships, drawings all preserved)
  - For-loop row expansion uses openpyxl in memory, then only the
    expanded worksheet XML is taken — everything else from template

Key fix: sharedStrings uses <t> tags, drawings use <a:t> tags.
These require SEPARATE regex patterns — using the wrong one silently
skips all rendering in that file.

Supports:
  - {{ variable }} in cells (via sharedStrings) and text boxes (via drawings)
  - {% for item in list %} ... {% endfor %} row expansion in cells
  - All images, drawings, text boxes, merged cells, styles fully preserved
"""

import argparse
import io
import json
import re
import sys
import zipfile
from copy import copy
from pathlib import Path

from jinja2 import Environment, Undefined
from openpyxl import load_workbook


# ----------------------------------------------------------------
# CLI
# ----------------------------------------------------------------
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generic XLSX generator")
    parser.add_argument("--template",        required=True)
    parser.add_argument("--output-filename", required=True)
    parser.add_argument("--context",         required=True)
    return parser.parse_args()


# ----------------------------------------------------------------
# Jinja2
# ----------------------------------------------------------------
def make_env() -> Environment:
    return Environment(undefined=Undefined)


def try_render(env: Environment, text: str, context: dict) -> str:
    """Render Jinja2, returning original text on any error."""
    if "{{" not in text and "{%" not in text:
        return text
    try:
        return env.from_string(text).render(context)
    except Exception:
        return text


# ----------------------------------------------------------------
# XML rendering — two separate functions for two different tag styles
# ----------------------------------------------------------------

def render_shared_strings(xml_bytes: bytes, context: dict, env: Environment) -> bytes:
    """
    Render {{ }} inside <t>...</t> tags in sharedStrings.xml.
    sharedStrings uses plain <t> (no namespace prefix).
    """
    try:
        xml = xml_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return xml_bytes

    def replace(m):
        attrs   = m.group(1)
        content = m.group(2)
        return f"<t{attrs}>{try_render(env, content, context)}</t>"

    xml = re.sub(r"<t([^>]*)>(.*?)</t>", replace, xml, flags=re.DOTALL)
    return xml.encode("utf-8")


def render_drawing_xml(xml_bytes: bytes, context: dict, env: Environment) -> bytes:
    """
    Render {{ }} inside <a:t>...</a:t> tags in drawing XML files.
    Drawing XML uses the 'a:' namespace prefix on text run elements.
    Using the wrong pattern (<t> instead of <a:t>) causes silent failure.
    """
    try:
        xml = xml_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return xml_bytes

    def replace(m):
        attrs   = m.group(1)
        content = m.group(2)
        return f"<a:t{attrs}>{try_render(env, content, context)}</a:t>"

    xml = re.sub(r"<a:t([^>]*)>(.*?)</a:t>", replace, xml, flags=re.DOTALL)
    return xml.encode("utf-8")


# ----------------------------------------------------------------
# Loop expansion — openpyxl processes rows, we only take cell XML
# ----------------------------------------------------------------

def has_loops(template_path: str) -> bool:
    """Check if the template contains {% for %} markers in cells."""
    with zipfile.ZipFile(template_path, "r") as z:
        # Check sharedStrings (where cell text lives)
        if "xl/sharedStrings.xml" in z.namelist():
            try:
                if "{%" in z.read("xl/sharedStrings.xml").decode("utf-8"):
                    return True
            except Exception:
                pass
        # Check worksheet XML for inline strings
        for name in z.namelist():
            if re.match(r"xl/worksheets/sheet\d+\.xml$", name):
                try:
                    if "{%" in z.read(name).decode("utf-8"):
                        return True
                except Exception:
                    pass
    return False


def render_cell_value(env: Environment, value, context: dict):
    if not isinstance(value, str):
        return value
    if "{{" not in value and "{%" not in value:
        return value
    try:
        rendered = env.from_string(value).render(context)
        stripped = rendered.strip()
        if stripped and stripped.lstrip("-").replace(".", "", 1).isdigit():
            return float(stripped) if "." in stripped else int(stripped)
        return rendered
    except Exception:
        return value


def find_loop_blocks(ws) -> list:
    for_re  = re.compile(r"\{%-?\s*for\s+(\w+)\s+in\s+(\w+)\s*-?%\}")
    end_re  = re.compile(r"\{%-?\s*endfor\s*-?%\}")
    blocks, pending = [], []

    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            fm = for_re.search(cell.value)
            if fm:
                pending.append((cell.row, fm.group(1), fm.group(2)))
            em = end_re.search(cell.value)
            if em and pending:
                for_row, var, list_key = pending.pop()
                blocks.append({"for_row": for_row, "end_row": cell.row,
                                "var": var, "list_key": list_key})

    blocks.sort(key=lambda b: b["for_row"], reverse=True)
    return blocks


def expand_loop(ws, block: dict, context: dict, env: Environment):
    for_row  = block["for_row"]
    end_row  = block["end_row"]
    var      = block["var"]
    list_key = block["list_key"]
    items    = context.get(list_key, [])
    if not isinstance(items, list):
        items = []

    body_rows = list(range(for_row, end_row + 1))
    num_body  = len(body_rows)
    max_col   = ws.max_column

    if not items:
        for _ in range(num_body):
            ws.delete_rows(for_row)
        return

    snapshot = []
    for r in body_rows:
        row_data = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row_data.append({
                "value":         cell.value,
                "font":          copy(cell.font)      if cell.has_style else None,
                "fill":          copy(cell.fill)      if cell.has_style else None,
                "border":        copy(cell.border)    if cell.has_style else None,
                "alignment":     copy(cell.alignment) if cell.has_style else None,
                "number_format": cell.number_format,
            })
        snapshot.append(row_data)

    extra = (len(items) - 1) * num_body
    if extra > 0:
        ws.insert_rows(end_row + 1, amount=extra)

    for item_idx, item in enumerate(items):
        item_ctx = dict(context)
        item_ctx[var] = item
        item_ctx["loop"] = {
            "index" : item_idx + 1,
            "index0" : item_idx,
            "first" : item_idx == 0,
            "last" : item_idx == len(items) - 1,
            "length" : len(items),
        }
        for body_idx, snap_row in enumerate(snapshot):
            tgt_row = for_row + item_idx * num_body + body_idx
            for col_idx, cd in enumerate(snap_row):
                tgt = ws.cell(row=tgt_row, column=col_idx + 1)
                tgt.value = render_cell_value(env, cd["value"], item_ctx)
                if cd["font"]:
                    tgt.font          = cd["font"]
                    tgt.fill          = cd["fill"]
                    tgt.border        = cd["border"]
                    tgt.alignment     = cd["alignment"]
                    tgt.number_format = cd["number_format"]


def expand_loops_to_xml(template_path: str, context: dict, env: Environment) -> dict:
    """
    Expand for-loops using openpyxl.
    Returns only the expanded worksheet XML files — nothing else.
    """
    wb = load_workbook(template_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for block in find_loop_blocks(ws):
            expand_loop(ws, block, context, env)
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                val = cell.value.strip()
                if re.match(r"^\{%-?\s*(for|endfor)\b", val):
                    cell.value = None
                    continue
                cell.value = render_cell_value(env, cell.value, context)

    buf = io.BytesIO()
    wb.save(buf)

    result = {}
    with zipfile.ZipFile(io.BytesIO(buf.getvalue()), "r") as z:
        for name in z.namelist():
            if re.match(r"xl/worksheets/sheet\d+\.xml$", name):
                result[name] = z.read(name)
    return result


# ----------------------------------------------------------------
# Main assembly
# ----------------------------------------------------------------

DRAWING_XML_RE = re.compile(r"xl/drawings/.*\.xml$", re.IGNORECASE)


def build_output(template_path: str, context: dict, env: Environment) -> bytes:
    """
    Assemble the final xlsx:
    - sharedStrings.xml  → rendered with <t> pattern
    - xl/drawings/*.xml  → rendered with <a:t> pattern
    - worksheet XML      → from openpyxl if loops present, else from template
    - Everything else    → verbatim from template
    """
    # Only invoke openpyxl if the template has for-loops
    expanded_sheets = {}
    if has_loops(template_path):
        expanded_sheets = expand_loops_to_xml(template_path, context, env)

    output_buf = io.BytesIO()

    with zipfile.ZipFile(template_path, "r") as tz, \
         zipfile.ZipFile(output_buf, "w", zipfile.ZIP_DEFLATED) as out:

        for name in tz.namelist():

            # Worksheet with expanded loops — use openpyxl version
            if name in expanded_sheets:
                out.writestr(name, expanded_sheets[name])

            # sharedStrings — render {{ }} with <t> pattern
            elif name == "xl/sharedStrings.xml":
                out.writestr(name, render_shared_strings(tz.read(name), context, env))

            # Drawing XML — render {{ }} with <a:t> pattern
            elif DRAWING_XML_RE.match(name):
                out.writestr(name, render_drawing_xml(tz.read(name), context, env))

            # Everything else — verbatim
            else:
                out.writestr(name, tz.read(name))

    return output_buf.getvalue()


# ----------------------------------------------------------------
# Entry point
# ----------------------------------------------------------------
def main() -> None:
    args = parse_args()

    try:
        context = json.loads(args.context)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON context — {e}", file=sys.stderr)
        sys.exit(1)

    base_dir      = Path(__file__).resolve().parent.parent
    template_path = str(base_dir / "document_templates" / args.template)
    output_dir    = base_dir / "public" / "cached_result"
    output_path   = output_dir / args.output_filename

    if not Path(template_path).exists():
        print(f"ERROR: Template not found at {template_path}", file=sys.stderr)
        sys.exit(1)

    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        env         = make_env()
        final_bytes = build_output(template_path, context, env)

        with open(str(output_path), "wb") as f:
            f.write(final_bytes)

        print(f"OK: Saved to {output_path}")

    except Exception as e:
        print(f"ERROR: Rendering failed — {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()